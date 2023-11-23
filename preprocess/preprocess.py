#!/usr/bin/env python3

import json
import openpyxl
import os
from pprint import pprint


def sheetToDictionary(sheet, title, desc, idx):
    return_dic = {}
    return_dic['title'] = title
    return_dic['desc'] = desc
    count = 0
    key = None

    search_list = []



    for i in range(1, sheet.max_row+1):
        for j in range(1, sheet.max_column+1):
            elem = sheet.cell( row=i+1, column=j ).value
            if elem != None:
                key = f"text{count}"
                if key in return_dic.keys():
                    return_dic[key].append(elem)
                else:
                    return_dic[key] = []
                    return_dic[key].append(elem)
                    search_list.append({
                        "id": f"subitem_{title.split()[0]}_{idx}_{count}",
                        "sstring": f"{title} -> {elem}"
                        })

        if key and len(return_dic[key]) != 3:
            if len(return_dic[key]) == 2:
                return_dic[key].append('')
            if len(return_dic[key]) == 1:
                return_dic[key].append('')
                return_dic[key].append('')

        count += 1

    return return_dic, search_list

def deletePreviousJson(fileNames):
    for name in fileNames:
        fileName = f"{name}.json"
        if os.path.exists(fileName):
            os.remove(fileName)


def getPartialName(wb_sheet_names, fullName):
    for name in wb_sheet_names:
        if fullName.split()[0] == name.split()[0]:
            return name
    return None


def XlToJson(xlName):
    wb = openpyxl.load_workbook(xlName)
    wb_sheet_names = wb.sheetnames
    class_with_sname = {
            "Data Source Links":["Government Agency Data","International and US Agency Data","Microdata Surveys","Replication Data","Regional Data","Dataverses"],

            "Research Tool Links":["Publication Searching Sites","AI Research Tools","Plagiarism Checker","List of Predatory Publishers","Paraphrasing Tools","Tutorials","Articles and Learning Materials","Training Packages","Software Download Links"],

            "Community":["Forum"],

            "Featured Research":["Nobel Winning Papers","Beginner Friendly Papers","Past Monographs and Assignments","ESC Conducted Researches","Systematic Reviews and Meta Analyses","Economics Department Researches"]
            }

    search = []
    deletePreviousJson(class_with_sname.keys())
    
    for key in class_with_sname.keys():
        sheet_data = []
        for idx, fullName in enumerate(class_with_sname[key]):
            partialName = getPartialName(wb_sheet_names, fullName)
            if partialName:
                sheetName = wb[partialName]
                return_dic, return_serach_list = sheetToDictionary(sheetName, fullName, '', idx)
                search += return_serach_list
                sheet_data.append(return_dic)
            else:
                print("error")

        with open(f"{key}.json", 'a') as of:
            json.dump(sheet_data, of)


    with open(f"search.json", 'w') as of:
        json.dump(search, of)

    


if __name__ == "__main__":
    XlToJson('./target.xlsx')


