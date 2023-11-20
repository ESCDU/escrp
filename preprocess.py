#!/usr/bin/env python3

import json
import openpyxl
from pprint import pprint


def sheetToDictionary(sheet, title, desc):
    return_dic = {}
    return_dic['title'] = title
    return_dic['desc'] = desc
    count = 0
    key = None



    for i in range(1, sheet.max_row+1):
        for j in range(1, sheet.max_column+1):
            elem = sheet.cell( row=i+1, column=j ).value
            if elem != None:
                key = f"text{count}"
                if key in return_dic.keys():
                    return_dic[key].append(elem)
                else:
                    return_dic[key] = []
                    return_dic[key].append(elem)

        if key and len(return_dic[key]) != 3:
            if len(return_dic[key]) == 2:
                return_dic[key].append('')
            if len(return_dic[key]) == 1:
                return_dic[key].append('')
                return_dic[key].append('')

        count += 1

    return return_dic


def getPartialName(wb_sheet_names, fullName):
    for name in wb_sheet_names:
        if fullName.split()[0] == name.split()[0]:
            return name
    return None


def XlToJson(xlName):
    wb = openpyxl.load_workbook(xlName)
    wb_sheet_names = wb.sheetnames
    class_with_sname = {
            "Data Source Links":["Government Agency Data","International and US Agency Data","Microdata Surveys","Replication Data","Regional Data","Dataverses"],

            "Research Tool Links":["Publication Searching Sites","AI Research Tools","Plagiarism Checker","List of Predatory Publishers","Paraphrasing Tools","Tutorials","Articles and Learning Materials","Training Packages","Software Download Links"],

            "Community":["Forum"],

            "Featured Research":["Nobel Winning Papers","Beginner Friendly Papers","Past Monographs and Assignments","ESC Conducted Researches","Systematic Reviews and Meta Analyses","Economics Department Researches"]
            }

    # pprint(wb_sheet_names)
    
    for key in class_with_sname.keys():
        sheet_data = []
        for fullName in class_with_sname[key]:
            partialName = getPartialName(wb_sheet_names, fullName)
            if partialName:
                sheetName = wb[partialName]
                sheet_data.append(sheetToDictionary(sheetName, fullName, ''))
            else:
                print("error")

        with open(f"{key}.json", 'a') as of:
            json.dump(sheet_data, of)

    
XlToJson('./target.xlsx')
